from typing import Pattern
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.alignment import Alignment
import pandas as pd
df = pd.read_csv('TempCol2022012513200154_SUM.csv')
writer = pd.ExcelWriter('TempCol2022012513200154_SUM.xlsx')
df.to_excel(writer, index=False)
writer.save()

wb = load_workbook('TempCol2022012513200154_SUM.xlsx')
ws = wb.active
# ws.print_options.horizontalCentered = True
# ws.print_options.verticalCentered = True
ws.insert_rows(1)
# ws.insert_rows(2)
# ws.delete_rows(1)

ws.merge_cells("A1:M1")
# ws.unmerge_cells("a1:m1")
# sheet["a1"]

ws['a1'] = 'Summary'
ws['a2'] = 'Collection'
ws['b2'] = 'Category'
ws['c2'] = 'BaseLine'
ws['d2'] = 'Total Deviation'
ws['e2'] = 'Actual Deviation'
ws['f2'] = 'Total Count of Checked Parameters'
ws['g2'] = 'Total Deviation %'
ws['h2'] = 'Actual Deviation %'
ws['i2'] = 'MO Summary'
ws['j2'] = 'Node Summary'
ws['k2'] = 'Parameter Summary'
ws['l2'] = 'Audit Summary'
ws['m2'] = 'SS Status'

ws['a1'].alignment = Alignment(horizontal="center")
ws['a1'].fill = PatternFill("solid", start_color="948A54")
ws['b2'].fill = PatternFill("solid", start_color="C4BD97")
ws['c2'].fill = PatternFill("solid", start_color="C4BD97")
ws['d2'].fill = PatternFill("solid", start_color="C4BD97")
ws['e2'].fill = PatternFill("solid", start_color="C4BD97")
ws['f2'].fill = PatternFill("solid", start_color="C4BD97")
ws['g2'].fill = PatternFill("solid", start_color="C4BD97")
ws['h2'].fill = PatternFill("solid", start_color="C4BD97")
ws['i2'].fill = PatternFill("solid", start_color="C4BD97")
ws['j2'].fill = PatternFill("solid", start_color="C4BD97")
ws['k2'].fill = PatternFill("solid", start_color="C4BD97")
ws['l2'].fill = PatternFill("solid", start_color="C4BD97")
ws['m2'].fill = PatternFill("solid", start_color="C4BD97")
wb.save('TempCol2022012513200154_SUM_header.xlsx')

# TempCol2022012513200154_SUM
